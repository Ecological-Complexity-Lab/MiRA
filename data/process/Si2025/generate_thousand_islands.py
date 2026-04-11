"""
Generate thousand_islands.json — plant-frugivore multilayer network
across 22 land-bridge islands at Thousand Island Lake, China.

Sources (Zhu et al. 2025, PNAS, doi:10.1073/pnas.2415846122):
  islands_attribute.xlsx            : island area (ha) and isolation (m)
  Bird_trait.xlsx                   : bird body mass, HWI, migrant status
  Plant_individual_frugivory_islands.xlsx : camera-trap frugivory events
  Interaction_metanetwork.xlsx      : meta-network edge roles (c, z, role)
"""

import json
import openpyxl
from collections import defaultdict

DATA_DIR = "."

# ---------------------------------------------------------------------------
# 1. Island attributes  (area in ha, isolation in m)
# ---------------------------------------------------------------------------

island_attrs = {}   # island_id -> {area, isolation}
wb = openpyxl.load_workbook(f"{DATA_DIR}/islands_attribute.xlsx",
                             read_only=True, data_only=True)
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    island_attrs[row[0]] = {"area_ha": round(row[1], 2),
                             "isolation_m": round(row[2], 2)}

ISLAND_ORDER = sorted(island_attrs, key=lambda x: int(x.replace("Island", "")))

# ---------------------------------------------------------------------------
# 2. Bird traits
# ---------------------------------------------------------------------------

bird_traits = {}    # bird_id -> {body_mass, HWI, migrant_status}
wb = openpyxl.load_workbook(f"{DATA_DIR}/Bird_trait.xlsx",
                             read_only=True, data_only=True)
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    bird_traits[row[0]] = {
        "body_mass_g":    round(row[1], 2),
        "HWI":            round(row[2], 2),
        "migrant_status": row[3],
    }

# ---------------------------------------------------------------------------
# 3. Meta-network interaction roles  (c, z, role per plant-bird pair)
# ---------------------------------------------------------------------------

meta_roles = {}     # (plant, bird) -> {c_value, z_value, role, compartment_group}
wb = openpyxl.load_workbook(f"{DATA_DIR}/Interaction_metanetwork.xlsx",
                             read_only=True, data_only=True)
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    plant, bird = row[0], row[1]
    meta_roles[(plant, bird)] = {
        "c_value":          round(row[3], 4) if row[3] is not None else None,
        "z_value":          round(row[4], 4) if row[4] is not None else None,
        "role":             row[5],
        "compartment_group": row[6],
    }

# ---------------------------------------------------------------------------
# 4. Frugivory interactions — aggregate frequency per (island, plant, bird)
# ---------------------------------------------------------------------------

raw_edges = defaultdict(int)
wb = openpyxl.load_workbook(f"{DATA_DIR}/Plant_individual_frugivory_islands.xlsx",
                             read_only=True, data_only=True)
for row in list(wb.active.iter_rows(values_only=True))[1:]:
    island, plant, _, bird, freq = row
    raw_edges[(island, plant, bird)] += freq

print(f"Aggregated intralayer edges: {len(raw_edges)}")

# Collect all state-nodes present
state_node_set = set()
for (isl, pl, bi) in raw_edges:
    state_node_set.add((pl, isl))
    state_node_set.add((bi, isl))

all_plants = sorted(set(pl for (isl, pl, bi) in raw_edges))
all_birds  = sorted(set(bi for (isl, pl, bi) in raw_edges))
print(f"Plants: {len(all_plants)}, Birds: {len(all_birds)}")

# ---------------------------------------------------------------------------
# 5. Interlayer links — Jaccard similarity of partner sets
# ---------------------------------------------------------------------------

partner_sets = defaultdict(lambda: defaultdict(set))
for (isl, pl, bi) in raw_edges:
    partner_sets[pl][isl].add(bi)    # plant's bird partners per island
    partner_sets[bi][isl].add(pl)    # bird's plant partners per island

interlayer_edges = []
for species, layer_partners in partner_sets.items():
    layers_present = [l for l in ISLAND_ORDER if l in layer_partners]
    if len(layers_present) < 2:
        continue
    for i in range(len(layers_present)):
        for j in range(i + 1, len(layers_present)):
            l1, l2 = layers_present[i], layers_present[j]
            s1, s2 = layer_partners[l1], layer_partners[l2]
            union = s1 | s2
            if not union:
                continue
            jaccard = round(len(s1 & s2) / len(union), 6)
            if jaccard == 0:
                continue
            interlayer_edges.append({
                "layer_from": l1,
                "node_from":  species,
                "layer_to":   l2,
                "node_to":    species,
                "weight":     jaccard,
            })

print(f"Interlayer edges: {len(interlayer_edges)}")

# ---------------------------------------------------------------------------
# 6. Build intralayer edge list with meta-network role attributes
# ---------------------------------------------------------------------------

intralayer_edges = []
for (isl, pl, bi), freq in raw_edges.items():
    edge = {
        "layer_from": isl,
        "node_from":  pl,
        "layer_to":   isl,
        "node_to":    bi,
        "weight":     int(freq),
    }
    role_info = meta_roles.get((pl, bi))
    if role_info:
        edge.update(role_info)
    intralayer_edges.append(edge)

# ---------------------------------------------------------------------------
# 7. Build JSON structure
# ---------------------------------------------------------------------------

layers = []
for i, isl in enumerate(ISLAND_ORDER):
    attrs = island_attrs[isl]
    layers.append({
        "layer_id":    i + 1,
        "layer_name":  isl,
        "bipartite":   True,
        "area_ha":     attrs["area_ha"],
        "isolation_m": attrs["isolation_m"],
    })

node_type = {}
for p in all_plants:
    node_type[p] = "plant"
for b in all_birds:
    node_type[b] = "bird"

nodes = []
for isl in ISLAND_ORDER:
    species_in_layer = sorted(
        (sp for (sp, layer) in state_node_set if layer == isl),
        key=lambda x: (node_type.get(x, "z"), x)
    )
    for sp in species_in_layer:
        ntype = node_type.get(sp, "unknown")
        entry = {
            "node_id":    sp,
            "layer_name": isl,
            "node_name":  sp,
            "node_type":  ntype,
        }
        if ntype == "bird":
            entry.update(bird_traits.get(sp, {}))
        nodes.append(entry)

state_nodes = [
    {"layer_name": layer, "node_name": sp}
    for (sp, layer) in sorted(state_node_set)
]

output = {
    "directed":            False,
    "directed_interlayer": False,
    "layers":              layers,
    "nodes":               nodes,
    "extended":            intralayer_edges + interlayer_edges,
    "state_nodes":         state_nodes,
}

# ---------------------------------------------------------------------------
# 8. Write output
# ---------------------------------------------------------------------------

OUT_PATH = "../../thousand_islands.json"
with open(OUT_PATH, "w") as f:
    json.dump(output, f, indent=2)

n_plants = len([n for n in nodes if n["node_type"] == "plant"])
n_birds  = len([n for n in nodes if n["node_type"] == "bird"])
roles_found = sum(1 for e in intralayer_edges if "role" in e)

print(f"\nWritten: {OUT_PATH}")
print(f"Layers:          {len(layers)} islands")
print(f"State-nodes:     {len(nodes)} ({n_plants} plant × island, {n_birds} bird × island)")
print(f"Intralayer:      {len(intralayer_edges)} edges (camera-trap visit frequency)")
print(f"  with role attr:{roles_found}/{len(intralayer_edges)}")
print(f"Interlayer:      {len(interlayer_edges)} edges (Jaccard partner similarity)")
print(f"Bird traits:     body_mass_g, HWI, migrant_status")
print(f"Link attrs:      weight, c_value, z_value, role, compartment_group")
print(f"Layer attrs:     area_ha, isolation_m")
